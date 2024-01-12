import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import csv

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

wb = load_workbook("Words.xlsx")
ws = wb['Sheet1']
max_row = ws.max_row

def add_word_action(new_word):  #code for adding new word to Excel file
    global max_row
    repeating_words = set(ws.cell(row=row, column=1).value for row in range(2, max_row + 1)) #if word already is in Excel, it is not added
    if new_word!="" and new_word not in repeating_words:
        new_word = new_word.capitalize() #words starts with capital letter
        ws.cell(row=max_row + 1, column=1, value=new_word)
        max_row += 1

        driver.get("https://www.deepl.com/en/translator")
        time.sleep(2)

        for row in range(2, max_row + 1):
            word = ws.cell(row=row, column=1).value
    
            find_input = driver.find_element(By.XPATH, "//div[@contenteditable='true' and @role='textbox']")
            find_input.clear()
            find_input.send_keys(new_word)
            time.sleep(2)

            find_output = driver.find_element(By.XPATH, "//*[@id='headlessui-tabs-panel-7']/div/div[1]/section/div/div[2]/div[3]/section/div[1]/d-textarea/div")
            translated_word = find_output.text
            ws.cell(row=max_row, column=2, value=translated_word)
            break
        if ws.cell(row=max_row, column=2).value == "":
            ws.delete_rows(max_row)
    else:
        print('You did not write a word or it is already in the list')

    new_headers = ['Word', 'Translated Word'] #headers to excel file
    if ws.cell(row=1, column=1).value is None:
        for col_index, header in enumerate(new_headers, start=1):
            ws.cell(row=1, column=col_index, value=header)


choose_option = input('New word/Find word/Work with text: ')

if choose_option == 'New word':
    new_word = input('Word: ')
    add_word_action(new_word)

elif choose_option == 'Find word':
    choose_language = input('In which language you will input a word? lv/eng: ')
    if choose_language == 'lv' or choose_language == 'eng':
        word_list = False
        if choose_language == 'eng':
            new_word = input('Word: ')
            for row in range(2, max_row + 1):
                word = ws.cell(row=row, column=1).value
                if word == new_word:
                    word_list = True
                    translated_word = ws.cell(row=row, column=2).value
                    print('Translation: ' + translated_word)
                    break
            if not word_list:
                print('Here is not this word, you can add it')
                choose = input('Do you want to add? ' )
                if choose == 'yes':
                    add_word_action(new_word)
        else:
            new_word = input('Word: ')
            for row in range(2, max_row + 1):
                word = ws.cell(row=row, column=2).value
                if word == new_word:
                    word_list = True
                    translated_word = ws.cell(row=row, column=1).value
                    print('Translation: ' + translated_word)
                    break
            if not word_list:
                print('Here is not this word')
    else:
        print('Here is not that option. Choose between lv and eng')

elif choose_option == 'Work with text':
    transl_dict = {}
    result_text = []
    word_count=0
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        word = ws.cell(row=row, column=1).value
        translation = ws.cell(row=row, column=2).value
        if word is not None and translation is not None:
            transl_dict[word.lower()] = translation
    input_text = input("Enter the text: ")
    split_words = input_text.split()
    for word in split_words:
        lower_word = word.lower() 
        if lower_word in transl_dict:
            result_text.append(f'{word} ({transl_dict[lower_word]})')
        else:
            result_text.append(word)
        word_count += 1  # after every 10th word, text goes on next row
        if word_count % 11 == 0:
            result_text.append('\n')
    translated_text = ' '.join(result_text)
    with open('text.csv', 'w', newline='', encoding='utf-8') as csv_file:
        csv_writer = csv.writer(csv_file)
        csv_writer.writerow([translated_text])


else :
    print('Error')
    print('Try one more time')

data = list(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2, values_only=True))  #words in exel file is sorted alphabetically first row and second row automaticly changes with cell from first column
sorted_data = sorted(data, key=lambda x: x[0])
ws.delete_rows(2, ws.max_row)
for row_index, row_data in enumerate(sorted_data, start=2):
    for col_index, value in enumerate(row_data, start=1):
        ws.cell(row=row_index, column=col_index, value=value)

wb.save("Words.xlsx")
wb.close()
driver.quit()